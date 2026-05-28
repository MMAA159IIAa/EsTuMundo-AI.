import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import sqlite3
from datetime import datetime
from tkcalendar import DateEntry
from dateutil.relativedelta import relativedelta
import csv
import os
import sys
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.chart import BarChart, Reference

BG_DARK      = "#0d0d0d"
BG_CARD      = "#1a1a1a"
BG_INPUT     = "#242424"
ACCENT       = "#e8a020"
ACCENT_DARK  = "#b87d10"
TEXT_PRIMARY = "#f0f0f0"
TEXT_MUTED   = "#888888"
SUCCESS      = "#2ecc71"
DANGER       = "#e74c3c"
BORDER       = "#2a2a2a"
FONT_TITLE   = ("Segoe UI", 22, "bold")
FONT_HEADER  = ("Segoe UI", 11, "bold")
FONT_BODY    = ("Segoe UI", 10)
FONT_SMALL   = ("Segoe UI", 9)

def obtener_ruta_db():
    if getattr(sys, 'frozen', False):
        base_path = os.path.join(os.environ['LOCALAPPDATA'], "TallerPro")
    else:
        base_path = os.path.dirname(os.path.abspath(__file__))
    if not os.path.exists(base_path):
        os.makedirs(base_path)
    return os.path.join(base_path, "taller_pro.db")

DB_NAME = obtener_ruta_db()

def conectar():
    return sqlite3.connect(DB_NAME)

def crear_tablas():
    conn = conectar()
    c = conn.cursor()
    c.execute("""CREATE TABLE IF NOT EXISTS registros (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nombre TEXT, telefono TEXT, correo TEXT,
        auto TEXT, numero_economico TEXT, placas TEXT,
        servicio TEXT, fecha TEXT, costo REAL DEFAULT 0,
        estatus TEXT DEFAULT 'Pendiente')""")
    c.execute("""CREATE TABLE IF NOT EXISTS finanzas (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        tipo TEXT, concepto TEXT, monto REAL, fecha TEXT, notas TEXT)""")
    c.execute("""CREATE TABLE IF NOT EXISTS config (
        clave TEXT PRIMARY KEY, valor TEXT)""")
    conn.commit()
    conn.close()

def boton(parent, texto, comando, color=ACCENT, texto_color=BG_DARK, w=14):
    return tk.Button(parent, text=texto, command=comando,
                     bg=color, fg=texto_color, font=FONT_HEADER,
                     bd=0, relief="flat", padx=12, pady=7,
                     cursor="hand2", width=w,
                     activebackground=ACCENT_DARK, activeforeground=BG_DARK)

def entrada(parent, width=30):
    e = tk.Entry(parent, bg=BG_INPUT, fg=TEXT_PRIMARY,
                 font=FONT_BODY, bd=0, relief="flat",
                 insertbackground=ACCENT, width=width)
    e.config(highlightthickness=1, highlightbackground=BORDER, highlightcolor=ACCENT)
    return e

class TallerPro(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("TALLER PRO - Sistema de Gestion")
        self.geometry("1400x750")
        self.configure(bg=BG_DARK)
        self.resizable(True, True)
        crear_tablas()
        self._build_ui()
        self.actualizar_dashboard()

    def _build_ui(self):
        self.sidebar = tk.Frame(self, bg=BG_CARD, width=200)
        self.sidebar.pack(side="left", fill="y")
        self.sidebar.pack_propagate(False)
        self.main = tk.Frame(self, bg=BG_DARK)
        self.main.pack(side="left", fill="both", expand=True)
        self._build_sidebar()
        self.frames = {}
        for Cls in (FrameDashboard, FrameClientes, FrameFinanzas, FrameReportes, FrameConfig):
            f = Cls(self.main, self)
            self.frames[Cls.__name__] = f
            f.place(relx=0, rely=0, relwidth=1, relheight=1)
        self.mostrar("FrameDashboard")

    def _build_sidebar(self):
        logo_frame = tk.Frame(self.sidebar, bg=BG_CARD, pady=20)
        logo_frame.pack(fill="x")
        tk.Label(logo_frame, text="TALLER PRO", bg=BG_CARD, fg=ACCENT,
                 font=("Segoe UI", 13, "bold")).pack()
        tk.Label(logo_frame, text="Sistema de Gestion", bg=BG_CARD,
                 fg=TEXT_MUTED, font=FONT_SMALL).pack()
        tk.Frame(self.sidebar, bg=BORDER, height=1).pack(fill="x", padx=15)
        nav_items = [
            ("Dashboard",    "FrameDashboard"),
            ("Clientes",     "FrameClientes"),
            ("Finanzas",     "FrameFinanzas"),
            ("Reportes",     "FrameReportes"),
            ("Configuracion","FrameConfig"),
        ]
        self.nav_buttons = {}
        for texto, frame in nav_items:
            btn = tk.Button(self.sidebar, text=texto, bg=BG_CARD,
                            fg=TEXT_PRIMARY, font=FONT_BODY, bd=0,
                            relief="flat", anchor="w", padx=20, pady=12,
                            cursor="hand2", width=22,
                            activebackground=ACCENT, activeforeground=BG_DARK,
                            command=lambda f=frame: self.mostrar(f))
            btn.pack(fill="x")
            self.nav_buttons[frame] = btn
        tk.Label(self.sidebar, text="v2.0 Pro", bg=BG_CARD,
                 fg=TEXT_MUTED, font=FONT_SMALL).pack(side="bottom", pady=10)

    def mostrar(self, nombre):
        for n, btn in self.nav_buttons.items():
            btn.config(bg=ACCENT if n == nombre else BG_CARD,
                       fg=BG_DARK if n == nombre else TEXT_PRIMARY)
        self.frames[nombre].tkraise()
        if hasattr(self.frames[nombre], "on_show"):
            self.frames[nombre].on_show()

    def actualizar_dashboard(self):
        if "FrameDashboard" in self.frames:
            self.frames["FrameDashboard"].actualizar()


class FrameDashboard(tk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent, bg=BG_DARK)
        self.app = app
        self._build()

    def _build(self):
        top = tk.Frame(self, bg=BG_DARK, pady=18, padx=25)
        top.pack(fill="x")
        tk.Label(top, text="Dashboard", bg=BG_DARK, fg=TEXT_PRIMARY,
                 font=FONT_TITLE).pack(side="left")
        tk.Label(top, text=datetime.now().strftime("%d/%m/%Y"),
                 bg=BG_DARK, fg=TEXT_MUTED, font=FONT_BODY).pack(side="right")
        self.kpi_frame = tk.Frame(self, bg=BG_DARK, padx=25)
        self.kpi_frame.pack(fill="x")
        self.kpis = {}
        kpi_defs = [
            ("clientes", "Clientes Totales", TEXT_PRIMARY),
            ("hoy",      "Servicios Hoy",    ACCENT),
            ("ingresos", "Ingresos del Mes", SUCCESS),
            ("egresos",  "Egresos del Mes",  DANGER),
            ("balance",  "Balance del Mes",  ACCENT),
            ("vencidos", "Serv. Vencidos",   DANGER),
        ]
        for key, titulo, color in kpi_defs:
            card = tk.Frame(self.kpi_frame, bg=BG_CARD, bd=0,
                            highlightthickness=1, highlightbackground=BORDER)
            card.pack(side="left", expand=True, fill="both", padx=6, pady=4)
            lbl_val = tk.Label(card, text="0", bg=BG_CARD, fg=color,
                               font=("Segoe UI", 20, "bold"))
            lbl_val.pack(pady=(14,2))
            tk.Label(card, text=titulo, bg=BG_CARD, fg=TEXT_MUTED,
                     font=FONT_SMALL).pack(pady=(2,14))
            self.kpis[key] = lbl_val
        mid = tk.Frame(self, bg=BG_DARK, padx=25, pady=10)
        mid.pack(fill="both", expand=True)
        tk.Label(mid, text="Ultimos Servicios", bg=BG_DARK,
                 fg=TEXT_PRIMARY, font=FONT_HEADER).pack(anchor="w")
        cols = ("Cliente","Auto","Servicio","Fecha","Costo","Estatus")
        self.tabla = ttk.Treeview(mid, columns=cols, show="headings", height=10)
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Treeview", background=BG_CARD, foreground=TEXT_PRIMARY,
                        fieldbackground=BG_CARD, rowheight=28, font=FONT_BODY)
        style.configure("Treeview.Heading", background=BG_INPUT,
                        foreground=ACCENT, font=FONT_HEADER, relief="flat")
        style.map("Treeview", background=[("selected", ACCENT)],
                  foreground=[("selected", BG_DARK)])
        for col in cols:
            self.tabla.heading(col, text=col)
            self.tabla.column(col, width=160)
        self.tabla.pack(fill="both", expand=True)

    def on_show(self):
        self.actualizar()

    def actualizar(self):
        conn = conectar()
        mes = datetime.now().strftime("%Y-%m")
        hoy = datetime.now().strftime("%Y-%m-%d")
        total_clientes = conn.execute("SELECT COUNT(*) FROM registros").fetchone()[0]
        servicios_hoy  = conn.execute("SELECT COUNT(*) FROM registros WHERE fecha=?", (hoy,)).fetchone()[0]
        ingresos = conn.execute("SELECT COALESCE(SUM(monto),0) FROM finanzas WHERE tipo='Ingreso' AND fecha LIKE ?", (mes+"%",)).fetchone()[0]
        egresos  = conn.execute("SELECT COALESCE(SUM(monto),0) FROM finanzas WHERE tipo='Egreso' AND fecha LIKE ?", (mes+"%",)).fetchone()[0]
        todos    = conn.execute("SELECT fecha FROM registros").fetchall()
        conn.close()
        vencidos = 0
        ahora = datetime.now()
        for (f,) in todos:
            try:
                if ahora >= datetime.strptime(f, "%Y-%m-%d") + relativedelta(months=6):
                    vencidos += 1
            except: pass
        self.kpis["clientes"].config(text=str(total_clientes))
        self.kpis["hoy"].config(text=str(servicios_hoy))
        self.kpis["ingresos"].config(text=f"${ingresos:,.0f}")
        self.kpis["egresos"].config(text=f"${egresos:,.0f}")
        self.kpis["balance"].config(text=f"${ingresos-egresos:,.0f}")
        self.kpis["vencidos"].config(text=str(vencidos))
        for row in self.tabla.get_children():
            self.tabla.delete(row)
        conn = conectar()
        filas = conn.execute("SELECT nombre,auto,servicio,fecha,costo,estatus FROM registros ORDER BY id DESC LIMIT 15").fetchall()
        conn.close()
        for fila in filas:
            self.tabla.insert("", "end", values=fila)


class FrameClientes(tk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent, bg=BG_DARK)
        self.app = app
        self._build()

    def _build(self):
        top = tk.Frame(self, bg=BG_DARK, padx=25, pady=14)
        top.pack(fill="x")
        tk.Label(top, text="Gestion de Clientes", bg=BG_DARK,
                 fg=TEXT_PRIMARY, font=FONT_TITLE).pack(side="left")
        btn_frame = tk.Frame(top, bg=BG_DARK)
        btn_frame.pack(side="right")
        boton(btn_frame, "+ Nuevo",        self.nuevo, w=10).pack(side="left", padx=4)
        boton(btn_frame, "Editar",         self.editar, color=BG_INPUT, texto_color=TEXT_PRIMARY, w=8).pack(side="left", padx=4)
        boton(btn_frame, "Borrar",         self.borrar, color=DANGER, texto_color=TEXT_PRIMARY, w=8).pack(side="left", padx=4)
        boton(btn_frame, "Enviar Correo",  self.enviar_correo, color="#2980b9", texto_color=TEXT_PRIMARY, w=12).pack(side="left", padx=4)
        boton(btn_frame, "Exportar",       self.exportar, color="#27ae60", texto_color=TEXT_PRIMARY, w=10).pack(side="left", padx=4)
        boton(btn_frame, "Importar Excel", self.importar, color="#8e44ad", texto_color=TEXT_PRIMARY, w=13).pack(side="left", padx=4)
        search = tk.Frame(self, bg=BG_DARK, padx=25)
        search.pack(fill="x")
        self.buscar_var = tk.StringVar()
        self.buscar_var.trace("w", lambda *a: self.cargar())
        e = tk.Entry(search, textvariable=self.buscar_var,
                     bg=BG_INPUT, fg=TEXT_PRIMARY, font=FONT_BODY,
                     bd=0, insertbackground=ACCENT, width=50)
        e.config(highlightthickness=1, highlightbackground=BORDER, highlightcolor=ACCENT)
        e.pack(side="left", padx=8, pady=8, ipady=5)
        frame_tabla = tk.Frame(self, bg=BG_DARK, padx=25)
        frame_tabla.pack(fill="both", expand=True)
        cols = ("ID","Cliente","Telefono","Correo","Auto","Num Eco","Placas","Servicio","Fecha","Costo","Estatus")
        self.tabla = ttk.Treeview(frame_tabla, columns=cols, show="headings")
        anchos = [40,140,100,160,120,90,90,160,90,80,90]
        for col, w in zip(cols, anchos):
            self.tabla.heading(col, text=col)
            self.tabla.column(col, width=w)
        self.tabla.tag_configure("vencido", background="#3d1515", foreground="#ff6b6b")
        self.tabla.tag_configure("listo",   background="#0d2b1a", foreground="#2ecc71")
        sb = ttk.Scrollbar(frame_tabla, orient="vertical", command=self.tabla.yview)
        self.tabla.configure(yscrollcommand=sb.set)
        self.tabla.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")
        self.cargar()

    def on_show(self):
        self.cargar()

    def cargar(self):
        texto = self.buscar_var.get().lower() if hasattr(self, "buscar_var") else ""
        for row in self.tabla.get_children():
            self.tabla.delete(row)
        conn = conectar()
        filas = conn.execute("SELECT * FROM registros ORDER BY id DESC").fetchall()
        conn.close()
        ahora = datetime.now()
        for fila in filas:
            if texto and not any(texto in str(v).lower() for v in fila[1:]):
                continue
            tag = ""
            try:
                venc = datetime.strptime(fila[8], "%Y-%m-%d") + relativedelta(months=6)
                if ahora >= venc:
                    tag = "vencido"
            except: pass
            if fila[10] == "Listo":
                tag = "listo"
            self.tabla.insert("", "end", iid=fila[0], values=fila, tags=(tag,))

    def nuevo(self):
        FormularioCliente(self, self.app)

    def editar(self):
        sel = self.tabla.focus()
        if not sel:
            messagebox.showwarning("Aviso", "Selecciona un registro primero")
            return
        datos = self.tabla.item(sel)["values"]
        FormularioCliente(self, self.app, datos=datos, id_reg=sel)

    def borrar(self):
        sel = self.tabla.focus()
        if not sel: return
        if messagebox.askyesno("Confirmar", "Borrar este registro?"):
            conn = conectar()
            conn.execute("DELETE FROM registros WHERE id=?", (sel,))
            conn.commit()
            conn.close()
            self.cargar()
            self.app.actualizar_dashboard()

    def enviar_correo(self):
        sel = self.tabla.focus()
        if not sel:
            messagebox.showwarning("Aviso", "Selecciona un cliente primero")
            return
        datos = self.tabla.item(sel)["values"]
        correo = datos[3]
        if not correo:
            messagebox.showwarning("Sin correo", "Este cliente no tiene correo registrado")
            return
        EnviarCorreo(self, datos[1], correo, datos[4], datos[7])

    def exportar(self):
        archivo = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Excel","*.xlsx")])
        if not archivo: return
        conn = conectar()
        filas = conn.execute("SELECT * FROM registros ORDER BY id").fetchall()
        conn.close()
        if not filas:
            messagebox.showwarning("Aviso", "No hay clientes registrados")
            return
        exportar_xlsx_clientes(filas, archivo)
        messagebox.showinfo("Exportado", f"{len(filas)} clientes exportados.\nGuardado en:\n{archivo}")

    def importar(self):
        archivo = filedialog.askopenfilename(
            filetypes=[("Excel","*.xlsx"),("CSV","*.csv")])
        if not archivo: return
        try:
            if archivo.endswith(".xlsx"):
                wb = openpyxl.load_workbook(archivo)
                ws = wb.active
                filas = list(ws.iter_rows(min_row=2, values_only=True))
            else:
                import csv as csvmod
                with open(archivo, encoding="utf-8") as f:
                    filas = list(csvmod.reader(f))[1:]
            conn = conectar()
            importados = 0
            for fila in filas:
                if len(fila) >= 4 and any(fila):
                    nombre   = str(fila[1] if len(fila) > 1 else fila[0] or "")
                    telefono = str(fila[2] if len(fila) > 2 else "")
                    correo   = str(fila[3] if len(fila) > 3 else "")
                    auto     = str(fila[4] if len(fila) > 4 else "")
                    num_eco  = str(fila[5] if len(fila) > 5 else "")
                    placas   = str(fila[6] if len(fila) > 6 else "")
                    servicio = str(fila[7] if len(fila) > 7 else "")
                    fecha    = str(fila[8] if len(fila) > 8 else datetime.now().strftime("%Y-%m-%d"))
                    costo    = float(fila[9]) if len(fila) > 9 and fila[9] else 0.0
                    estatus  = str(fila[10] if len(fila) > 10 else "Pendiente")
                    conn.execute("""INSERT INTO registros
                        (nombre,telefono,correo,auto,numero_economico,placas,servicio,fecha,costo,estatus)
                        VALUES (?,?,?,?,?,?,?,?,?,?)""",
                        (nombre,telefono,correo,auto,num_eco,placas,servicio,fecha,costo,estatus))
                    importados += 1
            conn.commit()
            conn.close()
            self.cargar()
            self.app.actualizar_dashboard()
            messagebox.showinfo("Importado", f"{importados} clientes importados correctamente")
        except Exception as ex:
            messagebox.showerror("Error", f"No se pudo importar:\n{ex}")


class FormularioCliente(tk.Toplevel):
    def __init__(self, parent, app, datos=None, id_reg=None):
        super().__init__(parent)
        self.app = app
        self.parent_frame = parent
        self.id_reg = id_reg
        self.title("Nuevo Registro" if not id_reg else "Editar Registro")
        self.geometry("480x620")
        self.configure(bg=BG_CARD)
        self.resizable(False, False)
        self._build(datos)

    def _build(self, datos):
        tk.Label(self, text="Datos del Servicio", bg=BG_CARD,
                 fg=ACCENT, font=FONT_TITLE).pack(pady=(20,10))
        campos = ["Cliente","Telefono","Correo","Auto","Num. Economico","Placas","Servicio"]
        self.entradas = []
        frame = tk.Frame(self, bg=BG_CARD, padx=30)
        frame.pack(fill="x")
        for campo in campos:
            tk.Label(frame, text=campo, bg=BG_CARD, fg=TEXT_MUTED, font=FONT_SMALL).pack(anchor="w", pady=(8,1))
            e = entrada(frame, width=42)
            e.pack(fill="x", ipady=5)
            self.entradas.append(e)
        row = tk.Frame(frame, bg=BG_CARD)
        row.pack(fill="x", pady=(8,0))
        col1 = tk.Frame(row, bg=BG_CARD)
        col1.pack(side="left", expand=True, fill="x", padx=(0,8))
        col2 = tk.Frame(row, bg=BG_CARD)
        col2.pack(side="left", expand=True, fill="x")
        tk.Label(col1, text="Fecha", bg=BG_CARD, fg=TEXT_MUTED, font=FONT_SMALL).pack(anchor="w")
        self.fecha = DateEntry(col1, date_pattern="yyyy-mm-dd", locale='es')
        self.fecha.pack(fill="x", ipady=4)
        tk.Label(col2, text="Costo ($)", bg=BG_CARD, fg=TEXT_MUTED, font=FONT_SMALL).pack(anchor="w")
        self.costo = entrada(col2, width=15)
        self.costo.pack(fill="x", ipady=5)
        tk.Label(frame, text="Estatus", bg=BG_CARD, fg=TEXT_MUTED, font=FONT_SMALL).pack(anchor="w", pady=(8,1))
        self.estatus = ttk.Combobox(frame, values=["Pendiente","En proceso","Listo","Entregado"], state="readonly", font=FONT_BODY)
        self.estatus.set("Pendiente")
        self.estatus.pack(fill="x", ipady=4)
        if datos:
            vals = list(datos)
            for i, e in enumerate(self.entradas):
                e.insert(0, str(vals[i+1]))
            try: self.fecha.set_date(vals[8])
            except: pass
            self.costo.insert(0, str(vals[9] or ""))
            self.estatus.set(vals[10] or "Pendiente")
        boton(self, "Guardar", self.guardar, w=20).pack(pady=20)

    def guardar(self):
        vals = [e.get() for e in self.entradas]
        fecha_val = self.fecha.get()
        try: costo_val = float(self.costo.get() or 0)
        except: costo_val = 0.0
        estatus_val = self.estatus.get()
        conn = conectar()
        if self.id_reg:
            conn.execute("""UPDATE registros SET nombre=?,telefono=?,correo=?,auto=?,
                numero_economico=?,placas=?,servicio=?,fecha=?,costo=?,estatus=?
                WHERE id=?""", (*vals, fecha_val, costo_val, estatus_val, self.id_reg))
        else:
            conn.execute("""INSERT INTO registros
                (nombre,telefono,correo,auto,numero_economico,placas,servicio,fecha,costo,estatus)
                VALUES (?,?,?,?,?,?,?,?,?,?)""", (*vals, fecha_val, costo_val, estatus_val))
        conn.commit()
        conn.close()
        self.parent_frame.cargar()
        self.app.actualizar_dashboard()
        self.destroy()


class FrameFinanzas(tk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent, bg=BG_DARK)
        self.app = app
        self._build()

    def _build(self):
        top = tk.Frame(self, bg=BG_DARK, padx=25, pady=14)
        top.pack(fill="x")
        tk.Label(top, text="Control de Finanzas", bg=BG_DARK,
                 fg=TEXT_PRIMARY, font=FONT_TITLE).pack(side="left")
        boton(top, "+ Agregar", self.nuevo, w=12).pack(side="right", padx=4)
        boton(top, "Borrar", self.borrar, color=DANGER, texto_color=TEXT_PRIMARY, w=10).pack(side="right", padx=4)
        resumen = tk.Frame(self, bg=BG_DARK, padx=25)
        resumen.pack(fill="x")
        self.lbl_ing = tk.Label(resumen, text="Ingresos: $0", bg=BG_CARD,
                                fg=SUCCESS, font=("Segoe UI",13,"bold"), padx=20, pady=10)
        self.lbl_ing.pack(side="left", padx=6)
        self.lbl_eg = tk.Label(resumen, text="Egresos: $0", bg=BG_CARD,
                               fg=DANGER, font=("Segoe UI",13,"bold"), padx=20, pady=10)
        self.lbl_eg.pack(side="left", padx=6)
        self.lbl_bal = tk.Label(resumen, text="Balance: $0", bg=BG_CARD,
                                fg=ACCENT, font=("Segoe UI",13,"bold"), padx=20, pady=10)
        self.lbl_bal.pack(side="left", padx=6)
        frame_t = tk.Frame(self, bg=BG_DARK, padx=25, pady=10)
        frame_t.pack(fill="both", expand=True)
        cols = ("ID","Tipo","Concepto","Monto","Fecha","Notas")
        self.tabla = ttk.Treeview(frame_t, columns=cols, show="headings")
        anchos = [40,80,200,100,100,300]
        for col, w in zip(cols, anchos):
            self.tabla.heading(col, text=col)
            self.tabla.column(col, width=w)
        self.tabla.tag_configure("ingreso", foreground=SUCCESS)
        self.tabla.tag_configure("egreso",  foreground=DANGER)
        sb = ttk.Scrollbar(frame_t, orient="vertical", command=self.tabla.yview)
        self.tabla.configure(yscrollcommand=sb.set)
        self.tabla.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")
        self.cargar()

    def on_show(self):
        self.cargar()

    def cargar(self):
        for row in self.tabla.get_children():
            self.tabla.delete(row)
        conn = conectar()
        filas = conn.execute("SELECT * FROM finanzas ORDER BY id DESC").fetchall()
        mes = datetime.now().strftime("%Y-%m")
        ing = conn.execute("SELECT COALESCE(SUM(monto),0) FROM finanzas WHERE tipo='Ingreso' AND fecha LIKE ?", (mes+"%",)).fetchone()[0]
        eg  = conn.execute("SELECT COALESCE(SUM(monto),0) FROM finanzas WHERE tipo='Egreso' AND fecha LIKE ?", (mes+"%",)).fetchone()[0]
        conn.close()
        for fila in filas:
            tag = "ingreso" if fila[1] == "Ingreso" else "egreso"
            self.tabla.insert("", "end", iid=fila[0], values=fila, tags=(tag,))
        self.lbl_ing.config(text=f"Ingresos del Mes: ${ing:,.2f}")
        self.lbl_eg.config(text=f"Egresos del Mes: ${eg:,.2f}")
        bal = ing - eg
        self.lbl_bal.config(text=f"Balance: ${bal:,.2f}", fg=SUCCESS if bal >= 0 else DANGER)
        self.app.actualizar_dashboard()

    def nuevo(self):
        FormularioFinanza(self)

    def borrar(self):
        sel = self.tabla.focus()
        if not sel: return
        if messagebox.askyesno("Confirmar", "Borrar este movimiento?"):
            conn = conectar()
            conn.execute("DELETE FROM finanzas WHERE id=?", (sel,))
            conn.commit()
            conn.close()
            self.cargar()


class FormularioFinanza(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.parent_frame = parent
        self.title("Nuevo Movimiento")
        self.geometry("420x380")
        self.configure(bg=BG_CARD)
        self.resizable(False, False)
        self._build()

    def _build(self):
        tk.Label(self, text="Nuevo Movimiento", bg=BG_CARD,
                 fg=ACCENT, font=FONT_TITLE).pack(pady=(20,10))
        frame = tk.Frame(self, bg=BG_CARD, padx=30)
        frame.pack(fill="x")
        tk.Label(frame, text="Tipo", bg=BG_CARD, fg=TEXT_MUTED, font=FONT_SMALL).pack(anchor="w")
        self.tipo = ttk.Combobox(frame, values=["Ingreso","Egreso"], state="readonly", font=FONT_BODY)
        self.tipo.set("Ingreso")
        self.tipo.pack(fill="x", ipady=4)
        tk.Label(frame, text="Concepto", bg=BG_CARD, fg=TEXT_MUTED, font=FONT_SMALL).pack(anchor="w", pady=(8,1))
        self.concepto = entrada(frame, width=42)
        self.concepto.pack(fill="x", ipady=5)
        tk.Label(frame, text="Monto ($)", bg=BG_CARD, fg=TEXT_MUTED, font=FONT_SMALL).pack(anchor="w", pady=(8,1))
        self.monto = entrada(frame, width=42)
        self.monto.pack(fill="x", ipady=5)
        tk.Label(frame, text="Fecha", bg=BG_CARD, fg=TEXT_MUTED, font=FONT_SMALL).pack(anchor="w", pady=(8,1))
        self.fecha = DateEntry(frame, date_pattern="yyyy-mm-dd", locale='es')
        self.fecha.pack(fill="x", ipady=4)
        tk.Label(frame, text="Notas", bg=BG_CARD, fg=TEXT_MUTED, font=FONT_SMALL).pack(anchor="w", pady=(8,1))
        self.notas = entrada(frame, width=42)
        self.notas.pack(fill="x", ipady=5)
        boton(self, "Guardar", self.guardar, w=20).pack(pady=20)

    def guardar(self):
        try: monto = float(self.monto.get() or 0)
        except: monto = 0.0
        conn = conectar()
        conn.execute("INSERT INTO finanzas (tipo,concepto,monto,fecha,notas) VALUES (?,?,?,?,?)",
                     (self.tipo.get(), self.concepto.get(), monto, self.fecha.get(), self.notas.get()))
        conn.commit()
        conn.close()
        self.parent_frame.cargar()
        self.destroy()


class FrameReportes(tk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent, bg=BG_DARK)
        self.app = app
        self._build()

    def _build(self):
        top = tk.Frame(self, bg=BG_DARK, padx=25, pady=14)
        top.pack(fill="x")
        tk.Label(top, text="Reportes", bg=BG_DARK, fg=TEXT_PRIMARY, font=FONT_TITLE).pack(side="left")
        content = tk.Frame(self, bg=BG_DARK, padx=25)
        content.pack(fill="both", expand=True)
        cards = [
            ("Reporte Mensual de Finanzas", "Excel con ingresos, egresos y balance del mes", self.reporte_finanzas),
            ("Reporte de Clientes", "Lista completa de clientes con servicios y costos", self.reporte_clientes),
            ("Reporte de Servicios Vencidos", "Clientes con servicio vencido hace mas de 6 meses", self.reporte_vencidos),
        ]
        for titulo, desc, cmd in cards:
            card = tk.Frame(content, bg=BG_CARD, padx=20, pady=15,
                            highlightthickness=1, highlightbackground=BORDER)
            card.pack(fill="x", pady=8)
            tk.Label(card, text=titulo, bg=BG_CARD, fg=ACCENT, font=FONT_HEADER).pack(anchor="w")
            tk.Label(card, text=desc, bg=BG_CARD, fg=TEXT_MUTED, font=FONT_SMALL).pack(anchor="w", pady=4)
            boton(card, "Generar", cmd, w=14).pack(anchor="w", pady=6)

    def reporte_finanzas(self):
        archivo = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel","*.xlsx")])
        if not archivo: return
        conn = conectar()
        mes = datetime.now().strftime("%Y-%m")
        filas = conn.execute("SELECT * FROM finanzas WHERE fecha LIKE ? ORDER BY fecha", (mes+"%",)).fetchall()
        conn.close()
        generar_reporte_finanzas(filas, archivo, mes)
        messagebox.showinfo("Reporte generado", f"Guardado en:\n{archivo}")

    def reporte_clientes(self):
        archivo = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel","*.xlsx")])
        if not archivo: return
        conn = conectar()
        filas = conn.execute("SELECT * FROM registros ORDER BY nombre").fetchall()
        conn.close()
        exportar_xlsx_clientes(filas, archivo)
        messagebox.showinfo("Reporte generado", f"Guardado en:\n{archivo}")

    def reporte_vencidos(self):
        archivo = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel","*.xlsx")])
        if not archivo: return
        conn = conectar()
        filas = conn.execute("SELECT * FROM registros").fetchall()
        conn.close()
        ahora = datetime.now()
        vencidos = [f for f in filas if _esta_vencido(f[8], ahora)]
        exportar_xlsx_clientes(vencidos, archivo, titulo="Servicios Vencidos")
        messagebox.showinfo("Reporte generado", f"{len(vencidos)} servicios vencidos.\nGuardado en:\n{archivo}")


def _esta_vencido(fecha_str, ahora):
    try:
        return ahora >= datetime.strptime(fecha_str, "%Y-%m-%d") + relativedelta(months=6)
    except:
        return False


class FrameConfig(tk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent, bg=BG_DARK)
        self.app = app
        self._build()

    def _build(self):
        top = tk.Frame(self, bg=BG_DARK, padx=25, pady=14)
        top.pack(fill="x")
        tk.Label(top, text="Configuracion de Correo", bg=BG_DARK,
                 fg=TEXT_PRIMARY, font=FONT_TITLE).pack(side="left")
        frame = tk.Frame(self, bg=BG_CARD, padx=30, pady=20)
        frame.pack(fill="x", padx=25, pady=10)
        campos = [
            ("Correo SMTP (Gmail)", "smtp_email"),
            ("Contrasena de App",   "smtp_pass"),
            ("Nombre del Taller",   "taller_nombre"),
            ("Telefono del Taller", "taller_tel"),
        ]
        self.entradas = {}
        for label_txt, key in campos:
            tk.Label(frame, text=label_txt, bg=BG_CARD, fg=TEXT_MUTED, font=FONT_SMALL).pack(anchor="w", pady=(8,1))
            e = entrada(frame, width=50)
            e.pack(fill="x", ipady=5)
            self.entradas[key] = e
        boton(frame, "Guardar Config", self.guardar, w=20).pack(pady=15)
        self.cargar_config()

    def cargar_config(self):
        conn = conectar()
        for key, e in self.entradas.items():
            row = conn.execute("SELECT valor FROM config WHERE clave=?", (key,)).fetchone()
            if row:
                e.delete(0, "end")
                e.insert(0, row[0])
        conn.close()

    def guardar(self):
        conn = conectar()
        for key, e in self.entradas.items():
            conn.execute("INSERT OR REPLACE INTO config (clave,valor) VALUES (?,?)", (key, e.get()))
        conn.commit()
        conn.close()
        messagebox.showinfo("OK", "Configuracion guardada")


class EnviarCorreo(tk.Toplevel):
    def __init__(self, parent, nombre, correo, auto, servicio):
        super().__init__(parent)
        self.title("Enviar Correo")
        self.geometry("500x400")
        self.configure(bg=BG_CARD)
        self.nombre = nombre
        self.correo = correo
        self.auto = auto
        self.servicio = servicio
        self._build()

    def _build(self):
        tk.Label(self, text="Enviar Correo al Cliente", bg=BG_CARD,
                 fg=ACCENT, font=FONT_TITLE).pack(pady=(20,10))
        frame = tk.Frame(self, bg=BG_CARD, padx=30)
        frame.pack(fill="x")
        tk.Label(frame, text=f"Para: {self.correo}", bg=BG_CARD, fg=TEXT_MUTED, font=FONT_SMALL).pack(anchor="w")
        tk.Label(frame, text="Asunto", bg=BG_CARD, fg=TEXT_MUTED, font=FONT_SMALL).pack(anchor="w", pady=(8,1))
        self.asunto = entrada(frame, width=50)
        self.asunto.insert(0, f"Tu vehiculo {self.auto} esta listo")
        self.asunto.pack(fill="x", ipady=5)
        tk.Label(frame, text="Mensaje", bg=BG_CARD, fg=TEXT_MUTED, font=FONT_SMALL).pack(anchor="w", pady=(8,1))
        self.mensaje = tk.Text(frame, bg=BG_INPUT, fg=TEXT_PRIMARY, font=FONT_BODY, bd=0, height=8, insertbackground=ACCENT)
        self.mensaje.pack(fill="x")
        self.mensaje.insert("1.0", f"Estimado/a {self.nombre},\n\nSu vehiculo {self.auto} ya esta listo.\n\nServicio: {self.servicio}\n\nGracias por su preferencia.")
        boton(self, "Enviar", self.enviar, w=20).pack(pady=15)

    def enviar(self):
        conn = conectar()
        smtp_email = conn.execute("SELECT valor FROM config WHERE clave='smtp_email'").fetchone()
        smtp_pass  = conn.execute("SELECT valor FROM config WHERE clave='smtp_pass'").fetchone()
        conn.close()
        if not smtp_email or not smtp_pass:
            messagebox.showwarning("Config", "Configura el correo SMTP primero")
            return
        try:
            msg = MIMEMultipart()
            msg["From"]    = smtp_email[0]
            msg["To"]      = self.correo
            msg["Subject"] = self.asunto.get()
            msg.attach(MIMEText(self.mensaje.get("1.0","end"), "plain"))
            server = smtplib.SMTP("smtp.gmail.com", 587)
            server.starttls()
            server.login(smtp_email[0], smtp_pass[0])
            server.send_message(msg)
            server.quit()
            messagebox.showinfo("OK", f"Correo enviado a {self.correo}")
            self.destroy()
        except Exception as ex:
            messagebox.showerror("Error", f"No se pudo enviar:\n{ex}")


def exportar_xlsx_clientes(filas, archivo, titulo="Clientes"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = titulo
    amarillo = PatternFill("solid", fgColor="E8A020")
    bold_b   = Font(bold=True, color="000000")
    center   = Alignment(horizontal="center", vertical="center")
    headers  = ["ID","Cliente","Telefono","Correo","Auto","Num Eco","Placas","Servicio","Fecha","Costo","Estatus"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = amarillo
        cell.font = bold_b
        cell.alignment = center
    for fila in filas:
        ws.append(list(fila))
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
    wb.save(archivo)


def generar_reporte_finanzas(filas, archivo, mes):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Finanzas {mes}"
    amarillo = PatternFill("solid", fgColor="E8A020")
    bold_b   = Font(bold=True, color="000000")
    center   = Alignment(horizontal="center")
    headers  = ["ID","Tipo","Concepto","Monto","Fecha","Notas"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = amarillo
        cell.font = bold_b
        cell.alignment = center
    ing = eg = 0
    for fila in filas:
        ws.append(list(fila))
        if fila[1] == "Ingreso": ing += fila[3]
        else: eg += fila[3]
    ws.append([])
    ws.append(["","","INGRESOS", ing])
    ws.append(["","","EGRESOS",  eg])
    ws.append(["","","BALANCE",  ing - eg])
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
    wb.save(archivo)


if __name__ == "__main__":
    app = TallerPro()
    app.mainloop()